from asyncio.windows_events import NULL
from traceback import print_tb
from tracemalloc import start
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from currency_web_scrapping import get_currency
import datetime
from duplicate_transcation import duplicate_to_main
from stock import stock

INCOME = 0
EXPENDITURE = 1
AUD_COL = 9
HKD_COL = 10
ACCOUNTS_COL = 7
DATE_COL = 15

##################################### HELPER FUNCTIONS #####################################

def fill_in_income_and_expenditure(start_row, start_col, type):
    end_col_of_income_and_expenditure = 5
    count = 1

    if type == INCOME:
        number_of_type = int(input("Number of INCOME: "))
    elif type == EXPENDITURE:
        number_of_type = int(input("Number of EXPENDITURE: "))
    for row in range(start_row, start_row + number_of_type):
        for col in range(start_col, end_col_of_income_and_expenditure, 2):
            char = get_column_letter(col)
            if col == 2:
                name_of_type = input(f"Type {count}: ")
                count += 1
                ws[char + str(row)].value = name_of_type
            elif col == 4:
                amount = float(input(f"Amount of {name_of_type}? "))
                ws[char + str(row)].value = amount

def get_currency_col_char_and_set_amount(currency, row, amount):
    char = get_column_letter(currency)
    ws[char + str(row)].value = amount

def set_amount(row):
    char = get_column_letter(ACCOUNTS_COL)
    while (ws[char + str(row)].value != None):
        if row == 12:
            break
        account_name = ws[char + str(row)].value
        amount = float(input(f"$ in {account_name}: "))
        if "AUD" in account_name:
            currency_char = get_column_letter(AUD_COL)
            ws[currency_char + str(row)].value = amount
        elif "HKD" in account_name:
            currency_char = get_column_letter(HKD_COL)
            ws[currency_char + str(row)].value = amount
        row += 1
    
############################################################################################

wb = load_workbook("2022 Personal Finance.xlsx")

month_num = input("Which month are you going to modify? ")
datetime_object = datetime.datetime.strptime(month_num, "%m")
sheet_name = datetime_object.strftime("%B")
ws = wb[sheet_name]

################################### INCOME & EXPENDITURE ###################################

start_row_of_income = 2
start_col_of_income = 2
start_row_of_expenditure = 14
start_col_of_expenditure = 2

fill_in_income_and_expenditure(start_row_of_income, start_col_of_income, INCOME)
fill_in_income_and_expenditure(start_row_of_expenditure, start_col_of_expenditure, EXPENDITURE)

############################################################################################

########################################## ASSETS ##########################################
AUD_COL = 9
HKD_COL = 10

# CASH
start_row_of_cash = 3
end_row_of_cash = 5

for row in range(start_row_of_cash, end_row_of_cash):
    if row == start_row_of_cash:
        amount_cash = float(input("AUD in Cash: "))
        char = get_column_letter(AUD_COL)
        ws[char + str(row)].value = amount_cash
    elif row == start_row_of_cash + 1:
        amount_cash = float(input("HKD in Cash: "))
        char = get_column_letter(HKD_COL)
        ws[char + str(row)].value = amount_cash
    
# ACCOUNTS
set_amount(6)

# Saving
set_amount(13)

############################################################################################

###########################################STOCK############################################
STOCK_NAME = 13
STOCK_QTY = 14
STOCK_PRICE = 15

start_row_of_stock = 3

number_of_stocks = int(input("Number of stocks?: "))
for row in range(start_row_of_stock, start_row_of_stock + number_of_stocks):
    stock_code = input("Stock code: ")
    stock_qty = float(input('Qty: '))

    code_char = get_column_letter(STOCK_NAME)
    qty_char = get_column_letter(STOCK_QTY)
    price_char = get_column_letter(STOCK_PRICE)

    ws[code_char + str(row)].value = stock_code
    ws[qty_char + str(row)].value = stock_qty
    ws[price_char + str(row)].value = stock(stock_code)

############################################################################################

###################################CURRENCY(WEB SCRAPPING)##################################
start_row_of_currency = 18
end_row_of_currency = 21
CURRENCY_TYPE_COL = 13
CURRENCY_EXCHANGE_AUD_COL = 14
CURRENCY_EXCHANGE_HKD_COL = 15
CURRENCY_EXCHANGE_USD_COL = 16

for row in range(start_row_of_currency, end_row_of_currency):
    if "AUD" in ws[get_column_letter(CURRENCY_TYPE_COL) + str(row)].value:
        ws[get_column_letter(CURRENCY_EXCHANGE_HKD_COL) + str(row)].value = get_currency("https://themoneyconverter.com/AUD/HKD")
        ws[get_column_letter(CURRENCY_EXCHANGE_USD_COL) + str(row)].value = get_currency("https://themoneyconverter.com/AUD/USD")
    elif "HKD" in ws[get_column_letter(CURRENCY_TYPE_COL) + str(row)].value:
        ws[get_column_letter(CURRENCY_EXCHANGE_AUD_COL) + str(row)].value = get_currency("https://themoneyconverter.com/HKD/AUD")
        ws[get_column_letter(CURRENCY_EXCHANGE_USD_COL) + str(row)].value = get_currency("https://themoneyconverter.com/HKD/USD")
    elif "USD" in ws[get_column_letter(CURRENCY_TYPE_COL) + str(row)].value:
        ws[get_column_letter(CURRENCY_EXCHANGE_AUD_COL) + str(row)].value = get_currency("https://themoneyconverter.com/USD/AUD")
        ws[get_column_letter(CURRENCY_EXCHANGE_HKD_COL) + str(row)].value = get_currency("https://themoneyconverter.com/USD/HKD")

today = datetime.date.today()
date = today.strftime("%d/%m/%Y")
char = get_column_letter(DATE_COL)
ws[char + str(16)].value = date

############################################################################################

################################### Duplicate Transcation ##################################
# transcation_wb = load_workbook("2022-01-01 _ 01-31.xlsx")
# transcation_ws = transcation_wb.active
# main_transcation_ws = wb["Transcation Details"]
# duplicate_to_main(transcation_ws, main_transcation_ws)

############################################################################################

print("Saved")
wb.save("2022 Personal Finance.xlsx")


                