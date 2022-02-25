from asyncio.windows_events import NULL
from json import load
from multiprocessing.dummy import current_process
from sre_constants import CATEGORY
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# PERIOD_COL = 1
# ACCOUNTS_COL = 2
# CATEGORY_COL = 3
# SUBCATEGORY_COL = 4
# NOTE_COL = 5
# AUD_COL = 6
# INCOME_EXPENSE_COL = 7
# DETAIL_COL = 8
# AMOUNT_COL = 9
# CURRENCY_COL = 10
# ACCOUNTS = 11

def copy_to_main(col, transcation_ws, main_ws):
    char = get_column_letter(col)
    row = 2
    current_cell = transcation_ws[char + str(row)].value
    while(current_cell != None):
        main_ws[char + str(row)].value = current_cell
        row += 1
        current_cell = transcation_ws[char + str(row)].value

def duplicate_to_main(transcation_ws, main_ws):
    for i in range(1, 12):
        copy_to_main(i, transcation_ws, main_ws)
