from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

sheet_name = ["Overview", "Transaction Details", "Template",
            "January", "February", "March", "April", "May", "June", "July", "August", "September",
            "October", "November", "December"]

wb = load_workbook("2022 Personal Finance.xlsx")
target = wb["Template"]

char = get_column_letter(7)

number_of_accounts = int(input("Number of accounts?(At most 6) "))
for row in range(6, 6 + number_of_accounts):
    print("Eg. Westpac Everyday Account(AUD), include (HKD) or (AUD) at the end")
    target[char + str(row)].value = input("Account Name: ")

number_of_saving = int(input("Number of saving?(At most 3): "))
for row in range(13, 13 + number_of_saving):
    print("Eg. Westpac Life(AUD), include (HKD) or (AUD) at the end")
    target[char + str(row)].value = input("Saving Name ")

for i in range(0, 12):
    wb.copy_worksheet(target)

for index, sheet in enumerate(wb):
    sheet.title = sheet_name[index]

print("Saved")
wb.save("2022 Personal Finance.xlsx")